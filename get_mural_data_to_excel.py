import requests
import json
from flask import Flask, request
import threading
import webbrowser
import time
import sys
import os
from urllib.parse import quote, urlparse, parse_qs
from datetime import datetime
import re
from bs4 import BeautifulSoup
import html
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

load_dotenv()

# === YOUR CREDENTIALS ===
CLIENT_ID = os.environ.get("MURAL_CLIENT_ID", "348d9c2f-0e87-4ca9-9ee5-10666fd0a14e")
CLIENT_SECRET = os.environ.get("MURAL_CLIENT_SECRET", "fae00b19f523ddb1cc9397ac5b9ea6e2facf295f20a96b3d7855b10ff1fc5e5dedb03972d34e4a19c62fa6a21856b41d0c7a578af8443858cc227ce46c8cb4ab")
REDIRECT_URI = os.environ.get("MURAL_REDIRECT_URI", "http://localhost:5000/callback")

app = Flask(__name__)
access_token = None
token_received = threading.Event()
flask_thread = None

# Clear any existing access token at startup
access_token = None

# Mural ID (extracted from your URLs)
MURAL_ID = "upwork4918.1764955053881"
MURAL_TITLE = "Test Salman-1"

# Patterns to filter out
FILTER_PATTERNS = [
    r"Very high scenario at 90%\(\s*2020\s*\)",
    r"Very high scenario at 90%\(\s*2040\s*\)",
    r"Very high scenario at 90%\s*\(\s*2060\s*\)",
    r"Very high scenario at 90%\s*\(\s*2080\s*\)",
    r"Very high scenario at 90%\s*\(\s*2099\s*\)",
    r"Very high scenario at 90% \(\s*2020\s*\)",
    r"Very high scenario at 90% \(\s*2040\s*\)",
    r"Very high scenario at 90% \(\s*2060\s*\)",
    r"Very high scenario at 90% \(\s*2080\s*\)",
    r"Very high scenario at 90% \(\s*2099\s*\)"
]


def filter_content(content):
    """Filter out unwanted patterns from content"""
    if not content:
        return content

    for pattern in FILTER_PATTERNS:
        content = re.sub(pattern, '', content, flags=re.IGNORECASE)

    # Clean up extra spaces that might result from removal
    content = re.sub(r'\s+', ' ', content).strip()

    return content


def get_mural_token_with_auth_code():
    """Get token using authorization code flow"""
    global access_token

    print("🔑 Starting OAuth flow...")
    print(f"📋 Using Redirect URI: {REDIRECT_URI}")

    # Use the EXACT scope names from Mural documentation
    scope = "rooms:read workspaces:read murals:read identity:read"
    print(f"🔐 Using Mural scopes: {scope}")

    # Properly encode all parameters
    auth_url = (
        f"https://app.mural.co/api/public/v1/authorization/oauth2/?"
        f"client_id={CLIENT_ID}&"
        f"redirect_uri={quote(REDIRECT_URI, safe='')}&"
        f"scope={quote(scope, safe='')}&"
        f"response_type=code"
    )

    print(f"🌐 Opening browser for Mural authorization...")

    # Try to open browser automatically
    try:
        webbrowser.open(auth_url)
    except Exception as e:
        print(f"⚠️  Could not open browser automatically: {e}")
        print(f"🔗 Please open this URL manually: {auth_url}")

    # Wait for the callback
    print("⏳ Waiting for authorization... (This will timeout after 120 seconds)")

    # Wait for token with timeout
    if token_received.wait(timeout=120):
        print("✅ Authorization successful!")
        return access_token
    else:
        print("❌ Authorization timeout - no response received within 120 seconds")
        return None


@app.route('/callback')
def oauth_callback():
    """Handle OAuth callback from Mural"""
    global access_token

    print("=" * 50)
    print("📥 Received callback from Mural!")
    print(f"🔍 Query parameters: {dict(request.args)}")
    print("=" * 50)

    # Get authorization code from query parameters
    auth_code = request.args.get('code')
    error = request.args.get('error')
    error_description = request.args.get('error_description', 'Unknown error')

    if error:
        print(f"❌ OAuth Error: {error}")
        print(f"📝 Error description: {error_description}")
        return f"<h1>Authorization Failed</h1><p>Error: {error}</p><p>Description: {error_description}</p>"

    if auth_code:
        print(f"✅ Received authorization code: {auth_code[:20]}...")

        # Exchange authorization code for access token
        token_url = "https://app.mural.co/api/public/v1/authorization/oauth2/token"

        # Prepare the data as form-urlencoded
        data = {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "redirect_uri": REDIRECT_URI,
            "code": auth_code,
            "grant_type": "authorization_code"
        }

        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }

        try:
            print("🔄 Exchanging authorization code for access token...")
            response = requests.post(token_url, headers=headers, data=data)

            print(f"📡 Token response status: {response.status_code}")

            if response.status_code == 200:
                token_data = response.json()
                access_token = token_data["access_token"]
                refresh_token = token_data.get("refresh_token")
                expires_in = token_data.get("expires_in", "Unknown")

                print("🎉 Successfully obtained access token!")
                print(f"   Access Token: {access_token[:50]}...")
                if refresh_token:
                    print(f"   Refresh Token: {refresh_token[:50]}...")
                print(f"   Expires in: {expires_in} seconds")

                # Signal that we have the token
                token_received.set()

                # Update .env file with new token
                try:
                    with open(".env", "r", encoding="utf-8") as f:
                        lines = f.readlines()
                    
                    with open(".env", "w", encoding="utf-8") as f:
                        found_access = False
                        found_refresh = False
                        for line in lines:
                            if line.startswith("MURAL_ACCESS_TOKEN="):
                                f.write(f"MURAL_ACCESS_TOKEN={access_token}\n")
                                found_access = True
                            elif line.startswith("MURAL_REFRESH_TOKEN=") and refresh_token:
                                f.write(f"MURAL_REFRESH_TOKEN={refresh_token}\n")
                                found_refresh = True
                            else:
                                f.write(line)
                        
                        if not found_access:
                            f.write(f"MURAL_ACCESS_TOKEN={access_token}\n")
                        if not found_refresh and refresh_token:
                            f.write(f"MURAL_REFRESH_TOKEN={refresh_token}\n")
                    print("✅ .env file updated with new tokens.")
                except Exception as e:
                    print(f"⚠️ Could not update .env: {e}")

                return """
                <!DOCTYPE html>
                <html>
                <head>
                    <title>✅ Authorization Successful</title>
                    <style>
                        body { font-family: Arial, sans-serif; margin: 40px; background: #f0f8f0; }
                        .success { background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
                    </style>
                </head>
                <body>
                    <div class="success">
                        <h1>✅ Authorization Successful!</h1>
                        <p>You can close this window and return to the application.</p>
                        <p>The script will continue automatically...</p>
                    </div>
                    <script>
                        setTimeout(function() { 
                            window.close(); 
                        }, 1000);
                    </script>
                </body>
                </html>
                """
            else:
                print(f"❌ Token exchange failed: {response.status_code}")
                print(f"   Response: {response.text}")
                token_received.set()
                return f"<h1>Token Exchange Failed</h1><p>Status: {response.status_code}</p><p>Response: {response.text}</p>"

        except Exception as e:
            print(f"❌ Error during token exchange: {e}")
            token_received.set()
            return f"<h1>Error</h1><p>{str(e)}</p>"
    else:
        print("❌ No authorization code received in callback")
        token_received.set()
        return "<h1>Error</h1><p>No authorization code received</p>"


@app.route('/')
def home():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Mural OAuth Server</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            .info { background: #f0f0f0; padding: 20px; border-radius: 10px; }
        </style>
    </head>
    <body>
        <div class="info">
            <h1>Mural OAuth Callback Server</h1>
            <p>Server is running and ready for OAuth callbacks.</p>
            <p>Endpoint: <code>/callback</p>
            <p>This window can be closed after authorization is complete.</p>
        </div>
    </body>
    </html>
    """


def clean_text_content(text):
    """Clean and normalize text content"""
    if text is None:
        return ""

    text = str(text)

    # Remove extra whitespace but keep all content
    text = re.sub(r'\s+', ' ', text).strip()

    # Decode HTML entities
    text = html.unescape(text)

    # Replace common HTML tags with spaces
    text = re.sub(r'<[^>]+>', ' ', text)

    # Remove non-breaking spaces
    text = text.replace('&nbsp;', ' ').replace('\xa0', ' ')

    # Clean up multiple spaces
    text = re.sub(r'[ \t]+', ' ', text).strip()

    return text


def extract_sticky_note_content(widget):
    """Extract ALL content from a sticky note widget"""
    content_items = []

    # Check ALL possible content fields
    content_fields = ['htmlText', 'text', 'title', 'content', 'description', 'label']

    for field in content_fields:
        if field in widget:
            field_value = widget[field]

            if isinstance(field_value, str) and field_value.strip():
                cleaned = clean_text_content(field_value)
                if cleaned:
                    # Filter out unwanted patterns
                    cleaned = filter_content(cleaned)
                    if cleaned:  # Only add if there's content after filtering
                        content_items.append({
                            'field': field,
                            'original': field_value[:500],  # Keep original for reference
                            'cleaned': cleaned
                        })
            elif isinstance(field_value, dict):
                # Check nested dictionary
                for key, value in field_value.items():
                    if isinstance(value, str) and value.strip():
                        cleaned = clean_text_content(value)
                        if cleaned:
                            # Filter out unwanted patterns
                            cleaned = filter_content(cleaned)
                            if cleaned:  # Only add if there's content after filtering
                                content_items.append({
                                    'field': f"{field}.{key}",
                                    'original': value[:500],
                                    'cleaned': cleaned
                                })

    return content_items


def get_widget_color(widget):
    """Get color information from widget - ENHANCED for all required colors"""
    style = widget.get('style', {})
    color_code = style.get('backgroundColor', '#FFFFFF')

    # Convert to string for comparison
    color_str = str(color_code).upper()

    # Enhanced color detection for all required colors from Excel
    color_mappings = [
        # Yellow - for Table 1 Column 1
        {"name": "Yellow",
         "codes": ["#FCF281", "#FFFF00", "255,255,0", "#FFEB3B", "#FFD700", "#FDD835", "#FFC107", "#FCF281FF"]},
        # Dark Red - for Table 1 Column 2
        {"name": "Dark Red",
         "codes": ["#BF0C0C", "191,12,12", "#D32F2F", "#C62828", "#B71C1C", "#FF0000", "#CC0000"]},
        # Orange - for Table 1 Column 3
        {"name": "Orange",
         "codes": ["#FFC061", "255,192,97", "#FF9800", "#F57C00", "#EF6C00", "#FFA500", "#FF8C00"]},
        # Green - for Table 2 RAPA
        {"name": "Green",
         "codes": ["#AAED92", "170,237,146", "#4CAF50", "#388E3C", "#2E7D32", "#66BB6A", "#81C784"]},
        # Blue - just in case
        {"name": "Blue",
         "codes": ["#9EDCFA", "158,220,250", "#2196F3", "#1976D2", "#1565C0"]},
        # White
        {"name": "White",
         "codes": ["#FFFFFF", "255,255,255", "#FFF", "#FAFAFA", "#FFFFFF00", "#FFFFFFFF"]},
    ]

    # First try exact matches
    for color_map in color_mappings:
        for code in color_map["codes"]:
            if code in color_str:
                return color_map["name"], color_code

    # If no exact match, try to parse RGBA/HEX
    if color_str.startswith("RGBA"):
        try:
            rgba = color_str.replace("RGBA", "").replace("(", "").replace(")", "").split(",")
            if len(rgba) >= 3:
                r = int(float(rgba[0].strip()))
                g = int(float(rgba[1].strip()))
                b = int(float(rgba[2].strip()))

                # Check for yellow
                if r > 200 and g > 200 and b < 150:
                    return "Yellow", color_code
                # Check for dark red
                elif r > 150 and g < 100 and b < 100:
                    return "Dark Red", color_code
                # Check for orange
                elif r > 200 and g > 100 and b < 100:
                    return "Orange", color_code
                # Check for green
                elif g > r and g > b:
                    return "Green", color_code
                # Check for blue
                elif b > r and b > g:
                    return "Blue", color_code
        except:
            pass

    # Check HEX colors
    if color_str.startswith("#"):
        hex_color = color_str.lstrip("#").upper()

        # Handle both 6-digit and 8-digit (with alpha) HEX
        if len(hex_color) in [6, 8]:
            try:
                # Take first 6 characters for RGB
                rgb_hex = hex_color[:6]
                r = int(rgb_hex[0:2], 16)
                g = int(rgb_hex[2:4], 16)
                b = int(rgb_hex[4:6], 16)

                # Check for yellow (fcf281 = RGB: 252, 242, 129)
                if r > 200 and g > 200 and b < 150:
                    return "Yellow", color_code
                # Check for dark red
                elif r > 150 and g < 100 and b < 100:
                    return "Dark Red", color_code
                # Check for orange
                elif r > 200 and g > 100 and b < 100:
                    return "Orange", color_code
                # Check for green
                elif g > r and g > b:
                    return "Green", color_code
            except:
                pass

    return "Other", color_code


def get_widget_position(widget):
    """Get position information from widget"""
    position = widget.get('position', {})

    # Try different possible position formats
    pos_x = 0
    pos_y = 0

    # Check for x coordinate
    if 'x' in position:
        pos_x = position['x']
    elif 'left' in position:
        pos_x = position['left']
    elif 'positionX' in position:
        pos_x = position['positionX']

    # Check for y coordinate
    if 'y' in position:
        pos_y = position['y']
    elif 'top' in position:
        pos_y = position['top']
    elif 'positionY' in position:
        pos_y = position['positionY']

    # Also check for nested position objects
    if isinstance(pos_x, dict):
        pos_x = pos_x.get('value', 0)
    if isinstance(pos_y, dict):
        pos_y = pos_y.get('value', 0)

    # Convert to numbers
    try:
        pos_x = float(pos_x) if pos_x else 0
        pos_y = float(pos_y) if pos_y else 0
    except:
        pos_x = 0
        pos_y = 0

    return pos_x, pos_y


def fetch_mural_widgets_with_pagination(token, mural_id):
    """Fetch ALL widgets from Mural with pagination support"""
    print(f"\n📋 Fetching ALL widgets from Mural (with pagination): {mural_id}")
    print("=" * 60)

    all_widgets = []
    next_token = None
    page_count = 0

    while True:
        page_count += 1

        # Build URL with pagination
        url = f"https://app.mural.co/api/public/v1/murals/{mural_id}/widgets"
        if next_token:
            url = f"{url}?next={next_token}"

        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json"
        }

        try:
            print(f"🌐 Fetching page {page_count}...")

            response = requests.get(url, headers=headers, timeout=30)

            if response.status_code == 200:
                data = response.json()

                # Get widgets from this page
                if isinstance(data, dict) and 'value' in data:
                    page_widgets = data['value']
                elif isinstance(data, list):
                    page_widgets = data
                else:
                    print(f"❌ Unexpected API response format on page {page_count}")
                    break

                all_widgets.extend(page_widgets)
                print(f"  ✅ Got {len(page_widgets)} widgets (total: {len(all_widgets)})")

                # Check for next page
                if 'next' in data and data['next']:
                    next_token = data['next']
                    print(f"  ↪️  More widgets available (next token: {next_token[:30]}...)")
                else:
                    print(f"  ✅ No more pages - fetched all widgets")
                    break
            else:
                print(f"❌ API request failed on page {page_count}: {response.status_code}")
                print(f"   Response: {response.text[:200]}")
                break

        except Exception as e:
            print(f"❌ Error fetching page {page_count}: {e}")
            import traceback
            traceback.print_exc()
            break

    print(f"\n📊 PAGINATION SUMMARY:")
    print(f"   • Pages fetched: {page_count}")
    print(f"   • Total widgets: {len(all_widgets)}")
    print("=" * 60)

    return all_widgets


def organize_sticky_notes_by_table_and_color(sticky_notes):
    """Organize sticky notes according to Excel table structure"""
    print("\n📊 Organizing sticky notes by table and color...")

    # Initialize table structure
    table1_data = {
        "yellow_notes": [],  # Column 1
        "dark_red_notes": [],  # Column 2
        "orange_notes": []  # Column 3
    }

    table2_data = {
        "green_notes": [],  # RAPA table - Adaptation Actions
        "blue_notes": []   # RAPA table - Assumptions (NEW)
    }

    other_notes = []  # Notes not matching required colors

    # Categorize notes by color
    for note in sticky_notes:
        color = note['color']

        if color == "Yellow":
            table1_data["yellow_notes"].append(note)
        elif color == "Dark Red":
            table1_data["dark_red_notes"].append(note)
        elif color == "Orange":
            table1_data["orange_notes"].append(note)
        elif color == "Green":
            table2_data["green_notes"].append(note)
        elif color == "Blue":  # NEW: Handle blue notes for assumptions
            table2_data["blue_notes"].append(note)
        else:
            other_notes.append(note)

    # Print summary
    print(f"\n📊 TABLE ORGANIZATION SUMMARY:")
    print(f"   Table 1 - Risks from climate change:")
    print(f"     • Column 1 (Yellow): {len(table1_data['yellow_notes'])} notes")
    print(f"     • Column 2 (Dark Red): {len(table1_data['dark_red_notes'])} notes")
    print(f"     • Column 3 (Orange): {len(table1_data['orange_notes'])} notes")
    print(f"   Table 2 - RAPA:")
    print(f"     • Adaptation Actions (Green): {len(table2_data['green_notes'])} notes")
    print(f"     • Assumptions (Blue): {len(table2_data['blue_notes'])} notes")  # UPDATED
    print(f"   Other colors: {len(other_notes)} notes")

    return table1_data, table2_data, other_notes


def apply_excel_formatting(filename):
    """Apply borders and formatting to Excel file"""
    print(f"\n🎨 Applying Excel formatting to {filename}...")

    try:
        # Load the workbook
        wb = load_workbook(filename)
        ws = wb.active

        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        # Apply borders to all cells with data
        max_row = ws.max_row
        max_col = ws.max_column

        # Apply thin borders to all cells with content
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value:  # Only apply borders to cells with content
                    cell.border = thin_border
                    # Apply text wrapping
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Apply thicker borders to header rows
        # Header rows are typically rows 1-5 (based on Excel structure)
        for row in range(1, 6):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.border = thick_border
                    # Make headers bold
                    cell.font = Font(bold=True)

        # Apply thicker border to Table 2 header (find it by looking for "Table 2: RAPA")
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
            for cell in row:
                if cell.value and "Table 2" in str(cell.value):
                    # Apply thick border to this row and next row (headers)
                    for col in range(1, max_col + 1):
                        ws.cell(row=cell.row, column=col).border = thick_border
                        ws.cell(row=cell.row, column=col).font = Font(bold=True)
                        if cell.row + 1 <= max_row:
                            ws.cell(row=cell.row + 1, column=col).border = thick_border
                            ws.cell(row=cell.row + 1, column=col).font = Font(bold=True)

        # Adjust column widths for better readability
        for col in range(1, max_col + 1):
            max_length = 0
            column = get_column_letter(col)

            # Find maximum length in column
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    # Account for line breaks
                    if '\n' in str(cell.value):
                        cell_length = max([len(line) for line in str(cell.value).split('\n')])
                    max_length = max(max_length, cell_length)

            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column].width = adjusted_width

        # Save the formatted workbook
        wb.save(filename)
        print(f"✅ Excel formatting applied successfully!")

    except Exception as e:
        print(f"⚠️  Could not apply Excel formatting: {e}")
        import traceback
        traceback.print_exc()


def create_json_for_report(table1_data, table2_data):
    """Create JSON structure for the report generation"""
    print("\n📝 Creating JSON structure for report...")

    structured_data = {
        'table1': {
            'title': 'Risks from climate change',
            'columns': [
                {
                    'header': 'Key current risks from extreme heat',
                    'content': [note['content'] for note in table1_data['yellow_notes']],
                    'color': 'Yellow'
                },
                {
                    'header': 'Key risks from a range of future climate scenarios, including the extreme heat event scenario (up to 50oC)',
                    'content': [note['content'] for note in table1_data['dark_red_notes']],
                    'color': 'Dark Red'
                },
                {
                    'header': 'Key adaptation actions to address different levels of risk (and their thresholds) for extreme heat',
                    'content': [note['content'] for note in table1_data['orange_notes']],
                    'color': 'Orange'
                }
            ]
        },
        'table2': {
            'title': 'RAPA (Rapid Adaptation Pathways Assessment)',
            'adaptation_actions': {  # UPDATED structure
                'content': [note['content'] for note in table2_data['green_notes']],
                'color': 'Green'
            },
            'assumptions': {  # NEW section
                'content': [note['content'] for note in table2_data['blue_notes']],
                'color': 'Blue'
            }
        }
    }

    # Save JSON file
    with open("mural_content_for_report.json", 'w', encoding='utf-8') as f:
        json.dump(structured_data, f, indent=2, ensure_ascii=False)

    print(f"✅ JSON file created: mural_content_for_report.json")
    return structured_data


def create_excel_output(table1_data, table2_data, mural_title):
    """Create Excel output matching the provided format"""
    print("\n💾 Creating Excel output...")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_title = "".join(c for c in mural_title if c.isalnum() or c in (' ', '_')).rstrip()
    excel_filename = f"Mural_Output_{safe_title}_{timestamp}.xlsx"

    # Create Excel writer
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Create DataFrame matching your Excel format

        # Start with Table 1 structure
        table1_rows = []

        # Get the maximum number of rows needed
        max_rows = max(
            len(table1_data['yellow_notes']),
            len(table1_data['dark_red_notes']),
            len(table1_data['orange_notes'])
        )

        # Add header rows first (as in your Excel)
        header_rows = [
            ["", "Appendix 6:", "", "", "", ""],
            ["[PLACEHOLDER]", "Table 1:", "", "", "", ""],
            ["TITLE", "Risks from climate change", "", "", "", ""],
            ["HEADERS COLUMN",
             "Key current risks from extreme heat",
             "Key risks from a range of future climate scenarios, including the extreme heat event scenario (up to 50oC)",
             "Key adaptation actions to address different levels of risk (and their thresholds) for extreme heat",
             "", ""]
        ]

        # Add empty row after headers (as in your Excel)
        header_rows.append(["", "", "", "", "", ""])

        # Create DataFrame for headers
        header_df = pd.DataFrame(header_rows)

        # Write headers to Excel
        header_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=0)

        # Add content rows for Table 1
        current_row = len(header_rows)

        # Add CONTENT (GREEN POST-ITS) row
        content_header_df = pd.DataFrame([["CONTENT (GREEN POST-ITS)", "", "", "", "", ""]])
        content_header_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=current_row)
        current_row += 1

        # Fill in content from sticky notes
        for i in range(max_rows):
            # Get content for each column
            yellow_content = table1_data['yellow_notes'][i]['content'] if i < len(table1_data['yellow_notes']) else ""
            dark_red_content = table1_data['dark_red_notes'][i]['content'] if i < len(
                table1_data['dark_red_notes']) else ""
            orange_content = table1_data['orange_notes'][i]['content'] if i < len(table1_data['orange_notes']) else ""

            # Create row
            row_df = pd.DataFrame([["", yellow_content, dark_red_content, orange_content, "", ""]])
            row_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=current_row)
            current_row += 1

        # Add empty rows (as in your Excel)
        for _ in range(3):
            empty_row_df = pd.DataFrame([["", "", "", "", "", ""]])
            empty_row_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=current_row)
            current_row += 1

        # Add Table 2 header
        table2_header_df = pd.DataFrame([["", "Table 2: RAPA", "", "", "", ""]])
        table2_header_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=current_row)
        current_row += 1

        # Add Table 2 sub-headers
        table2_subheaders_df = pd.DataFrame([["", "Adaptation Action", "30oC", "35oc", "Assumptions", "Uncertainties"]])
        table2_subheaders_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=current_row)
        current_row += 1

        # Add Table 2 content - match green adaptation actions with blue assumptions
        max_actions = max(len(table2_data['green_notes']), len(table2_data['blue_notes']))

        for i in range(max_actions):
            # Get adaptation action (green note)
            action = ""
            if i < len(table2_data['green_notes']):
                action = table2_data['green_notes'][i]['content']

            # Get assumption (blue note) - placed in Assumptions column (column E)
            assumption = ""
            if i < len(table2_data['blue_notes']):
                assumption = table2_data['blue_notes'][i]['content']

            # Create row with action in column B and assumption in column E
            row_df = pd.DataFrame([["", action, "", "", assumption, ""]])
            row_df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=current_row)
            current_row += 1

        print(f"✅ Excel file created: {excel_filename}")

        # Also create a summary sheet with raw data - UPDATED for blue notes
        summary_data = []
        for color, notes in [("Yellow", table1_data['yellow_notes']),
                             ("Dark Red", table1_data['dark_red_notes']),
                             ("Orange", table1_data['orange_notes']),
                             ("Green", table2_data['green_notes']),
                             ("Blue", table2_data['blue_notes'])]:  # ADDED blue notes
            for note in notes:
                summary_data.append({
                    "Color": color,
                    "Table": "Table 1" if color in ["Yellow", "Dark Red", "Orange"] else "Table 2",
                    "Column": "Assumptions" if color == "Blue" else "Adaptation Actions" if color == "Green" else color,
                    # ADDED column info
                    "Content": note['content'],
                    "Position_X": note['position_x'],
                    "Position_Y": note['position_y']
                })

        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Raw Data Summary', index=False)

    # Apply formatting to the Excel file
    apply_excel_formatting(excel_filename)

    return excel_filename


def process_all_widgets(widgets):
    """Process all widgets to extract sticky notes with proper data"""
    print(f"\n🔍 Processing {len(widgets)} widgets to extract sticky notes...")

    all_sticky_notes = []
    sticky_note_count = 0
    other_widgets = 0

    for widget_idx, widget in enumerate(widgets):
        widget_type = widget.get('type', 'unknown')
        widget_id = widget.get('id', '')

        # Get position
        pos_x, pos_y = get_widget_position(widget)

        # Get color
        color_name, color_code = get_widget_color(widget)

        # Check if this is a sticky note or text widget
        is_sticky_note = widget_type in ['sticky_note', 'stickyNote', 'sticky', 'text', 'textWidget', 'shape']

        # Process groups
        if widget_type == 'group':
            children = widget.get('children', [])

            for child_idx, child in enumerate(children):
                if isinstance(child, dict):
                    child_type = child.get('type', 'unknown')
                    child_id = child.get('id', '')

                    # Check if child is a sticky note
                    if child_type in ['sticky_note', 'stickyNote', 'sticky', 'text', 'textWidget', 'shape']:
                        sticky_note_count += 1

                        # Get child content
                        child_content_items = extract_sticky_note_content(child)

                        # Get child color
                        child_color_name, child_color_code = get_widget_color(child)

                        # Get child position
                        child_pos_x, child_pos_y = get_widget_position(child)

                        # Add each content item
                        for content_idx, content_item in enumerate(child_content_items):
                            all_sticky_notes.append({
                                'item_id': f"{widget_id[:8]}-child{child_idx}-{content_idx}",
                                'full_widget_id': child_id,
                                'widget_type': f"group_child_{child_type}",
                                'parent_id': widget_id,
                                'color': child_color_name,
                                'color_code': child_color_code,
                                'position_x': child_pos_x,
                                'position_y': child_pos_y,
                                'content': content_item['cleaned'],
                                'original_content': content_item['original'],
                                'source_field': content_item['field'],
                                'is_group_child': True
                            })

            # Also check group itself for content
            group_content = extract_sticky_note_content(widget)
            for content_item in group_content:
                all_sticky_notes.append({
                    'item_id': f"{widget_id[:8]}-group",
                    'full_widget_id': widget_id,
                    'widget_type': 'group',
                    'parent_id': None,
                    'color': color_name,
                    'color_code': color_code,
                    'position_x': pos_x,
                    'position_y': pos_y,
                    'content': content_item['cleaned'],
                    'original_content': content_item['original'],
                    'source_field': content_item['field'],
                    'is_group_child': False
                })

            continue

        # Process regular widgets
        if is_sticky_note:
            sticky_note_count += 1

            # Extract content
            content_items = extract_sticky_note_content(widget)

            # Add each content item
            for content_idx, content_item in enumerate(content_items):
                all_sticky_notes.append({
                    'item_id': f"{widget_id[:8]}-{content_idx}",
                    'full_widget_id': widget_id,
                    'widget_type': widget_type,
                    'parent_id': None,
                    'color': color_name,
                    'color_code': color_code,
                    'position_x': pos_x,
                    'position_y': pos_y,
                    'content': content_item['cleaned'],
                    'original_content': content_item['original'],
                    'source_field': content_item['field'],
                    'is_group_child': False
                })

            if content_items:
                content_preview = content_items[0]['cleaned'][:50] + "..." if len(content_items[0]['cleaned']) > 50 else \
                content_items[0]['cleaned']
                print(
                    f"  Widget {widget_idx + 1:03d}: {widget_type:15s} | {color_name:12s} | Pos({pos_x:.0f},{pos_y:.0f}) | {content_preview}")
            else:
                print(
                    f"  Widget {widget_idx + 1:03d}: {widget_type:15s} | {color_name:12s} | Pos({pos_x:.0f},{pos_y:.0f}) | [No content found]")
        else:
            other_widgets += 1
            # Check if it has content anyway
            content_items = extract_sticky_note_content(widget)
            if content_items:
                sticky_note_count += 1
                for content_idx, content_item in enumerate(content_items):
                    all_sticky_notes.append({
                        'item_id': f"{widget_id[:8]}-{content_idx}",
                        'full_widget_id': widget_id,
                        'widget_type': widget_type,
                        'parent_id': None,
                        'color': color_name,
                        'color_code': color_code,
                        'position_x': pos_x,
                        'position_y': pos_y,
                        'content': content_item['cleaned'],
                        'original_content': content_item['original'],
                        'source_field': content_item['field'],
                        'is_group_child': False
                    })
                content_preview = content_items[0]['cleaned'][:50] + "..." if len(content_items[0]['cleaned']) > 50 else \
                content_items[0]['cleaned']
                print(
                    f"  Widget {widget_idx + 1:03d}: {widget_type:15s} | {color_name:12s} | Pos({pos_x:.0f},{pos_y:.0f}) | {content_preview} (other type with content)")

    print(f"\n📊 WIDGET PROCESSING SUMMARY:")
    print(f"   • Total widgets processed: {len(widgets)}")
    print(f"   • Sticky notes found: {sticky_note_count}")
    print(f"   • Other widgets: {other_widgets}")
    print(f"   • Total content items extracted: {len(all_sticky_notes)}")

    return all_sticky_notes


# === MODIFIED SECTION IN create_text_report function ===
def create_text_report(table1_data, table2_data, other_notes, excel_filename):
    """Create a text report of the extraction"""
    print("\n📄 Creating text report...")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    report_content = f"""
{'=' * 80}
MURAL TO EXCEL EXTRACTION REPORT
{'=' * 80}
Generated: {timestamp}
Mural ID: {MURAL_ID}
Mural Title: {MURAL_TITLE}
Excel Output: {excel_filename}

{'=' * 80}
TABLE 1: RISKS FROM CLIMATE CHANGE
{'=' * 80}

Column 1 - Yellow Notes (Key current risks from extreme heat):
{'-' * 80}
"""

    for i, note in enumerate(table1_data['yellow_notes'], 1):
        report_content += f"\n{i}. {note['content']}\n"

    report_content += f"""
{'-' * 80}
Column 2 - Dark Red Notes (Key risks from future climate scenarios):
{'-' * 80}
"""

    for i, note in enumerate(table1_data['dark_red_notes'], 1):
        report_content += f"\n{i}. {note['content']}\n"

    report_content += f"""
{'-' * 80}
Column 3 - Orange Notes (Key adaptation actions):
{'-' * 80}
"""

    for i, note in enumerate(table1_data['orange_notes'], 1):
        report_content += f"\n{i}. {note['content']}\n"

    report_content += f"""
{'=' * 80}
TABLE 2: RAPA (Rapid Adaptation Pathways Assessment)
{'=' * 80}
Adaptation Actions (Green Notes):
{'-' * 80}
"""

    for i, note in enumerate(table2_data['green_notes'], 1):
        report_content += f"\n{i}. {note['content']}\n"

    report_content += f"""
{'-' * 80}
Assumptions (Blue Notes):
{'-' * 80}
"""

    for i, note in enumerate(table2_data['blue_notes'], 1):
        report_content += f"\n{i}. {note['content']}\n"

    if other_notes:
        report_content += f"""
{'=' * 80}
OTHER NOTES (Not categorized into tables)
{'=' * 80}
"""

        for i, note in enumerate(other_notes, 1):
            report_content += f"\n{i}. Color: {note['color']}\n   Content: {note['content']}\n"

    # Save report
    with open("mural_extraction_report.txt", 'w', encoding='utf-8') as f:
        f.write(report_content)

    print("✅ Text report created: mural_extraction_report.txt")


def main():
    """Main function to orchestrate the entire process"""
    global flask_thread

    print("=" * 70)
    print("🚀 MURAL TO EXCEL EXTRACTION TOOL")
    print("=" * 70)

    # Start Flask server in a separate thread
    flask_thread = threading.Thread(target=lambda: app.run(port=5000, debug=False, use_reloader=False))
    flask_thread.daemon = True
    flask_thread.start()
    print("✅ Flask server started on http://localhost:5000")
    time.sleep(2)  # Give Flask a moment to start

    # Get access token
    token = get_mural_token_with_auth_code()

    if not token:
        print("❌ Failed to obtain access token. Exiting.")
        return

    # Fetch all widgets from Mural
    widgets = fetch_mural_widgets_with_pagination(token, MURAL_ID)

    if not widgets:
        print("❌ No widgets found. Exiting.")
        return

    # Process widgets to extract sticky notes
    sticky_notes = process_all_widgets(widgets)

    if not sticky_notes:
        print("❌ No sticky note content found. Exiting.")
        return

    # Organize sticky notes by table and color
    table1_data, table2_data, other_notes = organize_sticky_notes_by_table_and_color(sticky_notes)

    # Create JSON for report
    structured_data = create_json_for_report(table1_data, table2_data)

    # Create Excel output
    excel_filename = create_excel_output(table1_data, table2_data, MURAL_TITLE)

    # Create a simple text report
    create_text_report(table1_data, table2_data, other_notes, excel_filename)

    print("\n" + "=" * 70)
    print("🎉 EXTRACTION COMPLETE!")
    print("=" * 70)
    print(f"📁 Output files created:")
    print(f"   • Excel file: {excel_filename}")
    print(f"   • JSON file: mural_content_for_report.json")
    print(f"   • Text report: mural_extraction_report.txt")
    print("\n📊 SUMMARY:")
    print(f"   • Table 1 - Risks from climate change: {sum(len(v) for v in table1_data.values())} notes")
    print(f"   • Table 2 - RAPA: {len(table2_data['green_notes'])} notes")
    print(f"   • Other notes: {len(other_notes)}")
    print("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ An error occurred: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)